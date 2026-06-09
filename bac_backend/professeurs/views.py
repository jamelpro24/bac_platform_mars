import os
import re
import tempfile
from difflib import SequenceMatcher

from django.contrib.auth import get_user_model
from django.db import transaction
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Professeur
from .serializers import ProfSerializer
from .windows_ocr import extract_pdf_text_with_windows_ocr


class ProfesseurViewSet(viewsets.ModelViewSet):
    serializer_class = ProfSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    # Matiere list (Arabic) used to validate/correct OCR.
    VALID_SPECIALITES = [
        "العربية",
        "الفرنسية",
        "الإنقليزية",
        "الأنقليزية",
        "الإعلامية",
        "التربية الإسلامية",
        "التاريخ والجغرافيا",
        "الرياضيات",
        "علوم الحياة والأرض",
        "العلوم الفيزيائية",
        "التربية البدنية",
        "التربية التقنية",
        "الألمانية",
        "الإسبانية",
        "التركية",
        "التربية المدنية",
        "إقنصاد",
        "الإقتصاد",
        "تصرف",
        "التربية التشكيلية",
        "التربية الموسيقية",
        "الهندسة الآلية",
        "التاريخ و الجغرافيا",
        "علوم الحياة و الأرض",
    ]

    SPECIALITE_ALIASES = {
        "العردية": "العربية",
        "العرية": "العربية",
        "الفرنسسية": "الفرنسية",
        "الأنقليزة": "الإنقليزية",
        "الأنقئيزية": "الإنقليزية",
        "الإعلاعية": "الإعلامية",
        "الإسلامية التربية": "التربية الإسلامية",
        "والجغرافي التا": "التاريخ والجغرافيا",
        "التاريخ و الجغرافيا": "التاريخ والجغرافيا",
        "التاريخ والجغرافي": "التاريخ والجغرافيا",
        "التقنية التردية": "التربية التقنية",
        "البدنية التردية": "التربية البدنية",
        "الموسيقية التربية": "التربية الموسيقية",
        "تاضيات": "الرياضيات",
        "اضيات أ": "الرياضيات",
        "اضيات": "الرياضيات",
        "الأرض و الحياة علو": "علوم الحياة والأرض",
        "علوم الحياة و الأرض": "علوم الحياة والأرض",
        "علوم الحياة الأرض": "علوم الحياة والأرض",
        "الفيزيائية العلو": "العلوم الفيزيائية",
        "الفيزيائية": "العلوم الفيزيائية",
        "اقتصد": "إقنصاد",
        "الاقتصاد": "الإقتصاد",
    }

    def get_queryset(self):
        qs = Professeur.objects.select_related("centre").all()
        if getattr(self.request.user, "role", "") == "admin":
            return qs.order_by("nom")
        return qs.filter(centre=self.request.user).order_by("nom")

    def perform_create(self, serializer):
        centre = self._resolve_centre()
        serializer.save(centre=centre)

    def perform_update(self, serializer):
        centre = self._resolve_centre()
        serializer.save(centre=centre)

    @action(detail=False, methods=["post"], url_path="import-pdf")
    def import_pdf(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "يرجى اختيار ملف PDF."}, status=status.HTTP_400_BAD_REQUEST)

        if not str(file.name or "").lower().endswith(".pdf"):
            return Response({"error": "الملف يجب أن يكون PDF."}, status=status.HTTP_400_BAD_REQUEST)

        payload = self._extract_pdf_payload(file)
        rows = self._parse_professeurs(payload.get("text", ""), ocr_lines=payload.get("lines", []))
        if not rows:
            return Response(
                {"error": "لم يتم العثور على بيانات قابلة للاستراد داخل الملف. تحقق من جودة المسح/اللغة."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = self._save_rows(rows)
        result["message"] = f"تم الاستيراد: إضافة {result['created']} وتحديث {result['updated']}."
        result["source"] = payload.get("source", "")
        return Response(result)

    @action(detail=False, methods=["post"], url_path="import-excel")
    def import_excel(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "يرجى اختيار ملف Excel."}, status=status.HTTP_400_BAD_REQUEST)

        filename = str(file.name or "").lower()
        if not filename.endswith((".xlsx", ".xlsm")):
            return Response({"error": "الملف يجب أن يكون Excel بصيغة .xlsx أو .xlsm."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = self._parse_excel_rows(file)
        except Exception as exc:
            return Response({"error": f"فشل قراءة ملف Excel: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response(
                {"error": "لم يتم العثور على بيانات قابلة للاستيراد داخل ملف Excel."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = self._save_rows(rows)
        result["message"] = f"تم الاستيراد: إضافة {result['created']} وتحديث {result['updated']}."
        result["source"] = "excel"
        return Response(result)

    @action(detail=False, methods=["post"], url_path="import-text")
    def import_text(self, request):
        text = str(request.data.get("text") or "").strip()
        if not text:
            return Response({"error": "يرجى لصق النص أولا."}, status=status.HTTP_400_BAD_REQUEST)

        rows = self._parse_text_lines(text)
        if not rows:
            return Response(
                {"error": "لم يتم العثور على أسطر قابلة للاستراد داخل النص الملصق."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = self._save_rows(rows)
        result["message"] = f"تم الاستيراد: إضافة {result['created']} وتحديث {result['updated']}."
        return Response(result)

    def _resolve_centre(self):
        # For directeur accounts, always use the logged-in user as centre.
        if getattr(self.request.user, "role", "") != "admin":
            return self.request.user

        centre_value = str(self.request.data.get("centre") or "").strip()
        if not centre_value:
            raise serializers.ValidationError("المركز إجباري لحساب المشرف.")

        User = get_user_model()
        centre = None
        if centre_value.isdigit():
            centre = User.objects.filter(id=int(centre_value), role="directeur").first()
        if not centre:
            centre = User.objects.filter(role="directeur", centre__iexact=centre_value).first()
        if not centre:
            centre = User.objects.filter(role="directeur", username__iexact=centre_value).first()
        if not centre:
            raise serializers.ValidationError("المركز غير موجود.")
        return centre

    def _extract_pdf_payload(self, file):
        # 1) Try embedded text if pypdf is installed.
        extracted_text = ""
        try:
            from pypdf import PdfReader  # type: ignore

            file.seek(0)
            reader = PdfReader(file)
            parts = [(page.extract_text() or "") for page in reader.pages]
            extracted_text = "\n".join(parts).strip()
        except Exception:
            extracted_text = ""

        if extracted_text:
            return {"text": extracted_text, "lines": [], "source": "pdf-text"}

        # 2) Fallback: Windows OCR (scanned PDFs).
        file.seek(0)
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
            tmp_pdf.write(file.read())
            tmp_path = tmp_pdf.name

        try:
            ocr_payload = extract_pdf_text_with_windows_ocr(tmp_path, language_tag="ar-SA")
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        ocr_text = str(ocr_payload.get("text") or "").strip()
        ocr_lines = [str(line).strip() for line in (ocr_payload.get("lines") or []) if str(line).strip()]
        return {"text": ocr_text, "lines": ocr_lines, "source": "windows-ocr"}

    def _parse_professeurs(self, text, ocr_lines=None):
        if ocr_lines:
            rows = self._parse_ocr_lines(ocr_lines)
            if rows:
                return rows
        rows = self._parse_text_lines(text)
        if rows:
            return rows
        return self._parse_lines_by_identifier(text)

    def _parse_ocr_lines(self, lines):
        normalized_lines = [" ".join(str(line).split()) for line in lines if str(line).strip()]
        if not normalized_lines:
            return []

        rows = self._parse_ocr_lines_by_headers(normalized_lines)
        if len(rows) >= 10:
            return rows

        fallback_rows = self._parse_ocr_lines_by_blocks(normalized_lines)
        if len(fallback_rows) > len(rows):
            return fallback_rows

        return rows

    def _parse_ocr_lines_by_headers(self, normalized_lines):
        institution_header = self._find_line_index(normalized_lines, lambda line: "المؤسسة" in line)
        specialite_header = self._find_line_index(
            normalized_lines,
            lambda line: "مادة" in line or "التدريس" in line or "الثدريس" in line,
        )
        names_header = self._find_line_index(
            normalized_lines,
            lambda line: ("اللقب" in line or "الاسم" in line) and ("الإسم" in line or "الاسم" in line),
        )
        first_identifier = self._find_line_index(normalized_lines, self._is_identifier_line)

        if None in (specialite_header, names_header, first_identifier):
            return []

        institutions = normalized_lines[(institution_header + 1 if institution_header is not None else 0) : specialite_header]
        specialites = normalized_lines[specialite_header + 1 : names_header]
        names = normalized_lines[names_header + 1 : first_identifier]
        identifiers = []

        for line in normalized_lines[first_identifier:]:
            if self._is_identifier_line(line):
                identifiers.append(self._normalize_identifier(line))
            elif identifiers:
                break

        institutions = [self._normalize_institution(line, rtl_reorder=True) for line in institutions]
        institutions = [line for line in institutions if line and not self._looks_like_noise_institution(line)]
        specialites = [self._normalize_specialite(line) for line in specialites]
        specialites = [line for line in specialites if line]
        names = [self._normalize_name(line, rtl_reorder=True) for line in names]
        names = [line for line in names if line and not self._looks_like_noise_name(line)]
        identifiers = [value for value in identifiers if value]

        count = min(len(institutions), len(specialites), len(names), len(identifiers))
        rows = []
        seen = set()
        for index in range(count):
            identifiant_unique = identifiers[index]
            if identifiant_unique in seen:
                continue
            seen.add(identifiant_unique)
            rows.append(
                {
                    "identifiant_unique": identifiant_unique,
                    "nom": names[index],
                    "specialite": specialites[index],
                    "institution": institutions[index],
                    "telephone": "",
                }
            )
        return rows

    def _parse_ocr_lines_by_blocks(self, normalized_lines):
        first_identifier = self._find_line_index(normalized_lines, self._is_identifier_line)
        if first_identifier is None:
            return []

        identifiers = []
        for line in normalized_lines[first_identifier:]:
            if self._is_identifier_line(line):
                value = self._normalize_identifier(line)
                if value:
                    identifiers.append(value)
            elif identifiers:
                break

        if not identifiers:
            return []

        expected_count = len(identifiers)
        before_ids = normalized_lines[:first_identifier]

        names, name_start = self._collect_last_valid(before_ids, expected_count, self._normalize_name, self._looks_like_noise_name)
        if len(names) != expected_count:
            return []

        specialite_source = before_ids[:name_start]
        specialites, specialite_start = self._collect_last_valid(
            specialite_source,
            expected_count,
            self._normalize_specialite,
            lambda value: False,
        )
        if len(specialites) != expected_count:
            return []

        institution_source = specialite_source[:specialite_start]
        institutions, _ = self._collect_last_valid(
            institution_source,
            expected_count,
            self._normalize_institution,
            self._looks_like_noise_institution,
        )
        if len(institutions) != expected_count:
            return []

        rows = []
        seen = set()
        for identifiant_unique, nom, specialite, institution in zip(identifiers, names, specialites, institutions):
            if identifiant_unique in seen:
                continue
            seen.add(identifiant_unique)
            rows.append(
                {
                    "identifiant_unique": identifiant_unique,
                    "nom": nom,
                    "specialite": specialite,
                    "institution": institution,
                    "telephone": "",
                }
            )

        return rows

    def _parse_text_lines(self, text):
        rows = []
        seen = set()
        for raw_line in (text or "").splitlines():
            # 1) Try tab-separated format first (Excel/text PDF export format)
            tab_parts = [p.strip() for p in raw_line.split("\t") if p.strip()]
            if len(tab_parts) >= 4:
                # Columns: [0]ع/ر, [1]المعرف الوحيد, [2]الاسم, [3]مادة التدريس, [4]المؤسسة, [5]ملاحظات
                identifiant_unique = self._normalize_identifier(tab_parts[1] if len(tab_parts) > 1 else "")
                nom = self._normalize_name(tab_parts[2] if len(tab_parts) > 2 else "")
                specialite = self._normalize_specialite(tab_parts[3] if len(tab_parts) > 3 else "")
                institution = self._normalize_institution(tab_parts[4] if len(tab_parts) > 4 else "")
                if identifiant_unique and nom and specialite:
                    if identifiant_unique not in seen:
                        seen.add(identifiant_unique)
                        rows.append({
                            "identifiant_unique": identifiant_unique,
                            "nom": nom,
                            "specialite": specialite,
                            "institution": institution or "",
                            "telephone": "",
                        })
                        continue

            # 2) Try splitting by 2+ spaces (PDF text extraction may preserve column spacing)
            space_parts = [p.strip() for p in re.split(r"\s{2,}", raw_line) if p.strip()]
            if len(space_parts) >= 4:
                identifiant_unique = self._normalize_identifier(space_parts[1] if len(space_parts) > 1 else "")
                nom = self._normalize_name(space_parts[2] if len(space_parts) > 2 else "")
                specialite = self._normalize_specialite(space_parts[3] if len(space_parts) > 3 else "")
                institution = self._normalize_institution(space_parts[4] if len(space_parts) > 4 else "")
                if identifiant_unique and nom and specialite:
                    if identifiant_unique not in seen:
                        seen.add(identifiant_unique)
                        rows.append({
                            "identifiant_unique": identifiant_unique,
                            "nom": nom,
                            "specialite": specialite,
                            "institution": institution or "",
                            "telephone": "",
                        })
                        continue

            # 3) Fallback: normalize whitespace and try old method
            line = " ".join((raw_line or "").split())
            if not line or self._is_header_or_noise(line):
                continue

            parts = [part.strip() for part in re.split(r"\t+|\s{2,}", line) if part.strip()]
            if len(parts) < 4:
                continue

            row = self._parse_text_parts(parts)
            if not row:
                continue
            if row["identifiant_unique"] in seen:
                continue
            seen.add(row["identifiant_unique"])
            rows.append(row)
        return rows

    def _parse_lines_by_identifier(self, text):
        rows = []
        seen = set()
        for raw_line in (text or "").splitlines():
            line = " ".join((raw_line or "").split())
            if not line or self._is_header_or_noise(line):
                continue

            # Match pure consecutive digits (8-12) first — avoids matching serial numbers
            match = re.search(r"\d{8,12}", line)
            if not match:
                # Fallback: try with separators (spaces, hyphens, dots, slashes)
                match = re.search(r"\d[\d\-./]{7,}\d", line)
            if not match:
                continue

            identifiant_unique = self._normalize_identifier(match.group(0))
            if not identifiant_unique or identifiant_unique in seen:
                continue

            before = line[: match.start()].strip()
            after = line[match.end() :].strip()
            candidate_text = " ".join(part for part in [before, after] if part)
            candidate_text = re.sub(r"^\d+\s+", "", candidate_text).strip()
            row = self._parse_identifier_line(candidate_text, identifiant_unique)
            if not row:
                continue

            seen.add(identifiant_unique)
            rows.append(row)
        return rows

    def _parse_identifier_line(self, text, identifiant_unique):
        tokens = [token for token in re.split(r"\s+", str(text).strip()) if token]
        if len(tokens) < 3:
            return None

        best = None
        for start in range(len(tokens)):
            for end in range(start + 1, min(len(tokens), start + 6) + 1):
                specialite = self._normalize_specialite(" ".join(tokens[start:end]))
                if specialite:
                    length = end - start
                    if best is None or length > best[0]:
                        best = (length, start, end, specialite)

        if not best:
            return None

        _, start, end, specialite = best
        left = " ".join(tokens[:start]).strip()
        right = " ".join(tokens[end:]).strip()

        left_is_institution = self._looks_like_institution_value(left)
        right_is_institution = self._looks_like_institution_value(right)

        if left and right:
            if left_is_institution and not right_is_institution:
                institution, nom = left, right
            elif right_is_institution and not left_is_institution:
                institution, nom = right, left
            else:
                institution, nom = left, right
        else:
            remaining = left or right
            parts = remaining.split()
            if len(parts) < 2:
                return None
            middle = max(1, len(parts) // 2)
            institution = " ".join(parts[:middle])
            nom = " ".join(parts[middle:])

        nom = self._normalize_name(nom)
        institution = self._normalize_institution(institution)

        if identifiant_unique and nom and specialite and institution:
            return {
                "identifiant_unique": identifiant_unique,
                "nom": nom,
                "specialite": specialite,
                "institution": institution,
                "telephone": "",
            }
        return None

    def _parse_text_parts(self, parts):
        # Arabic PDF table columns are read right-to-left:
        # المؤسسة | مادة التدريس | الاسم واللقب | المعرف الوحيد
        identifiant_unique = self._normalize_identifier(parts[-1])
        nom = self._normalize_name(parts[-2], rtl_reorder=True)
        specialite = self._normalize_specialite(parts[-3])
        institution = self._normalize_institution(parts[-4], rtl_reorder=True)

        if identifiant_unique and nom and specialite and institution:
            return {
                "identifiant_unique": identifiant_unique,
                "nom": nom,
                "specialite": specialite,
                "institution": institution,
                "telephone": "",
            }

        return None

    def _parse_excel_rows(self, file):
        from openpyxl import load_workbook

        file.seek(0)
        workbook = load_workbook(file, read_only=True, data_only=True)
        worksheet = workbook.active
        raw_rows = [
            ["" if cell is None else str(cell).strip() for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        workbook.close()

        raw_rows = [row for row in raw_rows if any(cell for cell in row)]
        if not raw_rows:
            return []

        header_index = self._find_excel_header_index(raw_rows)
        header_map = {}
        data_rows = raw_rows

        if header_index is not None:
            header_map = {
                self._excel_header_key(value): index
                for index, value in enumerate(raw_rows[header_index])
                if self._excel_header_key(value)
            }
            data_rows = raw_rows[header_index + 1 :]
        else:
            # Fallback: skip metadata rows and find first row with a valid identifier
            for i, row in enumerate(raw_rows):
                cells = [c for c in row if c]
                if any(self._normalize_identifier(c) for c in cells):
                    data_rows = raw_rows[i:]
                    break

        rows = []
        seen = set()
        for raw_row in data_rows:
            row = self._parse_excel_row(raw_row, header_map)
            if not row:
                continue
            if row["identifiant_unique"] in seen:
                continue
            seen.add(row["identifiant_unique"])
            rows.append(row)
        return rows

    def _find_excel_header_index(self, rows):
        # Search all rows (not just first 10) to handle ministry files with many metadata rows
        for index, row in enumerate(rows):
            joined = " ".join(str(c) for c in row)
            # Normalize tatweel and tashkeel for robust Arabic comparison
            joined_norm = re.sub(r"[\u0640\u064B-\u065F]", "", joined)
            joined_norm = joined_norm.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
            # Arabic header detection: look for identifier + name/speciality headers
            has_id = "المعرف" in joined_norm or "معرف" in joined_norm
            has_name = "الاسم" in joined_norm or "اللقب" in joined_norm
            has_specialite = "مادة" in joined_norm or "التدريس" in joined_norm or "الثدريس" in joined_norm
            if has_id and (has_name or has_specialite):
                return index
            # Latin header detection
            keys = {self._excel_header_key(value) for value in row}
            keys.discard("")
            if "identifiant_unique" in keys and ("nom" in keys or "specialite" in keys):
                return index
        return None

    def _excel_header_key(self, value):
        text = str(value or "").strip()
        # Normalize Arabic: remove tatweel, diacritics, shadda, alef variants
        text = re.sub(r"[\u0640\u064B-\u065F]", "", text)  # tatweel + tashkeel
        text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
        text = text.lower()
        text = re.sub(r"[\s_\-:/\\]+", " ", text)
        text = " ".join(text.split())
        compact = text.replace(" ", "")

        if any(token in compact for token in ["identifiant", "unique", "matricule", "cin", "id"]) or "المعرف" in compact or "معرف" in compact:
            return "identifiant_unique"
        if any(token in text for token in ["nom", "prenom", "prénom", "name"]) or "الاسم" in compact or "اللقب" in compact:
            return "nom"
        if any(token in text for token in ["specialite", "spécialité", "matiere", "matière"]) or "ماده" in compact or "مادة" in compact or "التدريس" in compact or "اختصاص" in compact:
            return "specialite"
        if any(token in text for token in ["institution", "etablissement", "établissement", "lycee", "lycée"]) or "المؤسسه" in compact or "المؤسسة" in compact:
            return "institution"
        if any(token in text for token in ["telephone", "téléphone", "phone", "tel"]):
            return "telephone"
        return ""

    def _parse_excel_row(self, raw_row, header_map):
        values = [str(value).strip() for value in raw_row]

        if header_map:
            def get_value(key):
                index = header_map.get(key)
                return values[index] if index is not None and index < len(values) else ""

            identifiant_unique = self._normalize_identifier(get_value("identifiant_unique"))
            nom = self._normalize_name(get_value("nom"))
            specialite = self._normalize_specialite(get_value("specialite")) or get_value("specialite")
            institution = self._normalize_institution(get_value("institution"))
            telephone = get_value("telephone")
        else:
            cells = [value for value in values if value]
            if len(cells) < 4:
                return None

            id_index = next((index for index, value in enumerate(cells) if self._normalize_identifier(value)), None)
            if id_index is None:
                return None

            identifiant_unique = self._normalize_identifier(cells[id_index])
            remaining = [value for index, value in enumerate(cells) if index != id_index]

            # Ministry format: after identifier comes nom, specialite, institution (id_index 0 or 1)
            if id_index <= 1 and len(remaining) >= 3:
                nom = self._normalize_name(remaining[0])
                specialite = self._normalize_specialite(remaining[1]) or remaining[1]
                institution = self._normalize_institution(remaining[2])
                telephone = remaining[3] if len(remaining) > 3 else ""
            elif id_index >= len(cells) - 2:
                institution = self._normalize_institution(remaining[0])
                specialite = self._normalize_specialite(remaining[1]) or remaining[1]
                nom = self._normalize_name(remaining[2])
                telephone = ""
            else:
                nom = self._normalize_name(remaining[0])
                specialite = self._normalize_specialite(remaining[1]) or remaining[1]
                institution = self._normalize_institution(remaining[2])
                telephone = remaining[3] if len(remaining) > 3 else ""

        if identifiant_unique and nom and specialite and institution:
            return {
                "identifiant_unique": identifiant_unique,
                "nom": nom,
                "specialite": specialite,
                "institution": institution,
                "telephone": telephone,
            }
        return None

    def _save_rows(self, rows):
        centre = self._resolve_centre()
        created = 0
        updated = 0

        with transaction.atomic():
            for row in rows:
                professeur = Professeur.objects.filter(centre=centre, identifiant_unique=row["identifiant_unique"]).first()
                if professeur:
                    professeur.nom = row["nom"]
                    professeur.specialite = row["specialite"]
                    professeur.institution = row["institution"]
                    professeur.telephone = row.get("telephone", "")
                    professeur.save(update_fields=["nom", "specialite", "institution", "telephone"])
                    updated += 1
                else:
                    Professeur.objects.create(
                        centre=centre,
                        identifiant_unique=row["identifiant_unique"],
                        nom=row["nom"],
                        specialite=row["specialite"],
                        institution=row["institution"],
                        telephone=row.get("telephone", ""),
                    )
                    created += 1

        return {
            "created": created,
            "updated": updated,
            "total": created + updated,
            "preview": rows[:10],
        }

    def _find_line_index(self, lines, predicate):
        for index, line in enumerate(lines):
            if predicate(line):
                return index
        return None

    def _collect_last_valid(self, lines, expected_count, normalize_fn, noise_fn):
        collected = []
        start_index = 0

        for index in range(len(lines) - 1, -1, -1):
            raw_value = lines[index]
            normalized = normalize_fn(raw_value, rtl_reorder=True) if normalize_fn in (self._normalize_name, self._normalize_institution) else normalize_fn(raw_value)
            if not normalized:
                continue
            if noise_fn(normalized):
                continue

            collected.append(normalized)
            start_index = index
            if len(collected) == expected_count:
                break

        collected.reverse()
        return collected, start_index

    def _is_identifier_line(self, line):
        digits = re.sub(r"\D", "", line)
        return digits.isdigit() and 8 <= len(digits) <= 12

    def _normalize_identifier(self, line):
        digits = re.sub(r"\D", "", str(line))
        if 8 <= len(digits) <= 12:
            return digits[:10].zfill(10)
        return ""

    def _normalize_name(self, line, rtl_reorder=False):
        cleaned = " ".join(str(line).replace('"', " ").replace(".", " ").split())
        if len(cleaned) < 4:
            return ""
        if re.fullmatch(r"[\d\s]+", cleaned):
            return ""
        if rtl_reorder:
            cleaned = self._reorder_rtl_words(cleaned)
        return cleaned

    def _normalize_institution(self, line, rtl_reorder=False):
        cleaned = " ".join(str(line).replace('"', " ").split())
        if len(cleaned) < 4:
            return ""
        if rtl_reorder:
            cleaned = self._reorder_rtl_words(cleaned)
        return cleaned

    def _reorder_rtl_words(self, value):
        parts = [part for part in str(value).split() if part]
        if len(parts) <= 1:
            return str(value).strip()
        return " ".join(reversed(parts))

    def _clean_specialite_text(self, line):
        cleaned = str(line).strip()
        cleaned = cleaned.replace('"', " ")
        cleaned = cleaned.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
        cleaned = cleaned.replace("ى", "ي")
        cleaned = re.sub(r"[^\u0600-\u06FF\s]", " ", cleaned)
        cleaned = " ".join(cleaned.split())

        # remove common header tokens
        for token in ("مادة", "التدريس", "الثدريس"):
            cleaned = cleaned.replace(token, "")

        return " ".join(cleaned.split())

    def _specialite_key(self, value):
        key = self._clean_specialite_text(value)
        key = key.replace("ة", "ه")
        key = key.replace("ؤ", "و").replace("ئ", "ي")
        key = key.replace(" ", "")
        return key

    def _normalize_specialite(self, line):
        cleaned = self._clean_specialite_text(line)
        if not cleaned:
            return ""

        cleaned_key = self._specialite_key(cleaned)

        for valid in self.VALID_SPECIALITES:
            if cleaned_key == self._specialite_key(valid):
                return valid

        for alias_key, alias_value in self.SPECIALITE_ALIASES.items():
            if cleaned_key == self._specialite_key(alias_key):
                return alias_value

        best_match = ""
        best_score = 0.0
        for valid in self.VALID_SPECIALITES:
            score = SequenceMatcher(None, cleaned_key, self._specialite_key(valid)).ratio()
            if score > best_score:
                best_score = score
                best_match = valid

        for alias_key, alias_value in self.SPECIALITE_ALIASES.items():
            score = SequenceMatcher(None, cleaned_key, self._specialite_key(alias_key)).ratio()
            if score > best_score:
                best_score = score
                best_match = alias_value

        if best_score >= 0.55:
            return best_match
        return ""

    def _looks_like_noise_name(self, line):
        return any(token in line for token in ["الوحيد", "المرف", "مادة", "المؤسسة"])

    def _looks_like_noise_institution(self, line):
        return any(token in line for token in ["المركز", "المؤسسة", "وزارة", "الجمهورية"])

    def _looks_like_institution_value(self, line):
        lowered = str(line or "").lower()
        return any(
            token in lowered
            for token in [
                "معهد",
                "مدرسة",
                "إعدادية",
                "اعدادية",
                "ثانوية",
                "ابتدائية",
                "lycee",
                "lycée",
                "ecole",
                "école",
                "college",
                "collège",
            ]
        )

    def _is_header_or_noise(self, line):
        lowered = line.lower()
        keywords = [
            "ministere",
            "liste",
            "professeurs",
            "nom",
            "prenom",
            "page ",
            "الجمهورية",
            "وزارة",
            "قائمة",
            "الاساتذة",
            "الأستاذة",
            "مادة التدريس",
            "الاسم",
            "المعرف",
            "المؤسسة",
        ]
        return any(keyword in lowered for keyword in keywords)